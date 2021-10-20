import vtk
from vtk.util import numpy_support
import sys
import os
import yaml
import numpy as np
from sklearn.decomposition import PCA
from openpyxl import Workbook
from collections import OrderedDict
from typing import List


def read_surfaces(ref_dir, def_dir):
    rsurfs = []
    for fname in sorted(os.listdir(ref_dir)):
        if '.stl' in fname.lower():
            reader = vtk.vtkSTLReader()
        elif ".vtk" in fname.lower():
            reader = vtk.vtkPolyDataReader()
        elif ".vtp" in fname.lower():
            reader = vtk.vtkXMLPolyDataReader()
        else:
            continue
        reader.SetFileName(
            str(os.path.normpath(ref_dir + os.sep + fname)))
        reader.Update()
        triangles = vtk.vtkTriangleFilter()
        triangles.SetInputConnection(reader.GetOutputPort())
        triangles.Update()
        rsurfs.append(triangles.GetOutput())
    dsurfs = []
    cell_ids = []
    for fname in sorted(os.listdir(def_dir)):
        if '.stl' in fname.lower():
            reader = vtk.vtkSTLReader()
        elif ".vtk" in fname.lower():
            reader = vtk.vtkPolyDataReader()
        elif ".vtp" in fname.lower():
            reader = vtk.vtkXMLPolyDataReader()
        else:
            continue
        reader.SetFileName(
            str(os.path.normpath(def_dir + os.sep + fname)))
        reader.Update()
        triangles = vtk.vtkTriangleFilter()
        triangles.SetInputConnection(reader.GetOutputPort())
        triangles.Update()
        dsurfs.append(triangles.GetOutput())
        cell_ids.append(os.path.basename(fname).replace('.stl', ''))
    return rsurfs, dsurfs, cell_ids


def get_cell_orientation(surf):
    x = numpy_support.vtk_to_numpy(surf.GetPoints().GetData())
    pca = PCA(n_components=3, svd_solver='full')
    pca.fit(x)
    return pca.components_[0]


def get_strain(rsurfs: List, dsurfs: List, cell_ids: List[int]) -> dict:
    rvolumes = []
    dvolumes = []
    rcentroids = []
    rsurface_areas = []
    dsurface_areas = []
    surface_area_strains = []
    cell_dirs = []
    vstrains = []
    strains = []
    principal_strain_1 = []
    principal_strain_3 = []
    principal_strain_1_direction = []
    principal_strain_3_direction = []
    for r, d in zip(rsurfs, dsurfs):
        mass_props = vtk.vtkMassProperties()
        mass_props.SetInputData(r)
        mass_props.Update()
        rvol = mass_props.GetVolume()
        rsurfarea = mass_props.GetSurfaceArea()
        celldir = get_cell_orientation(r)
        cell_dirs.append(celldir)
        mass_props.SetInputData(d)
        dvol = mass_props.GetVolume()
        dsurfarea = mass_props.GetSurfaceArea()
        rvolumes.append(rvol)
        dvolumes.append(dvol)
        rsurface_areas.append(rsurfarea)
        dsurface_areas.append(dsurfarea)
        center_of_mass = vtk.vtkCenterOfMass()
        center_of_mass.SetInputData(r)
        center_of_mass.Update()
        rcentroids.append(center_of_mass.GetCenter())
        # volumetric strains
        vstrains.append(dvol / rvol - 1.0)
        # surface area strains
        surface_area_strains.append(dsurfarea / rsurfarea - 1.0)

        iterative_closest_point = vtk.vtkIterativeClosestPointTransform()
        rcopy = vtk.vtkPolyData()
        dcopy = vtk.vtkPolyData()
        rcopy.DeepCopy(r)
        dcopy.DeepCopy(d)
        iterative_closest_point.SetSource(rcopy)
        iterative_closest_point.SetTarget(dcopy)
        iterative_closest_point.GetLandmarkTransform().SetModeToAffine()
        iterative_closest_point.SetMaximumMeanDistance(0.001)
        iterative_closest_point.SetCheckMeanDistance(1)
        iterative_closest_point.SetMaximumNumberOfIterations(5000)
        iterative_closest_point.StartByMatchingCentroidsOn()
        iterative_closest_point.Update()

        deformation_gradient = np.zeros((3, 3), float)
        for j in range(3):
            for k in range(3):
                deformation_gradient[j, k] = iterative_closest_point.GetMatrix().GetElement(j, k)
        green_lagrange_strain = 0.5 * (np.dot(deformation_gradient.T, deformation_gradient) - np.eye(3))
        principal_strains, principal_strain_directions = np.linalg.eigh(green_lagrange_strain)
        principal_strain_1.append(principal_strains[2])
        principal_strain_3.append(principal_strains[0])
        principal_strain_1_direction.append(principal_strain_directions[:, 2])
        principal_strain_3_direction.append(principal_strain_directions[:, 0])
        strains.append(green_lagrange_strain)
    results = OrderedDict({"Volumetric Strains": vstrains,
                           "Normalized Surface Area Change": surface_area_strains,
                           "Strains": strains,
                           "1st Principal Strains": principal_strain_1,
                           "3rd Principal Strains": principal_strain_3,
                           "1st Principal Strain Directions": principal_strain_1_direction,
                           "3rd Principal Strain Directions": principal_strain_3_direction,
                           "Reference Volumes": rvolumes,
                           "Reference Surface Areas": rsurface_areas,
                           "Deformed Volumes": dvolumes,
                           "Deformed Surface Areas": dsurface_areas,
                           "Reference Cell Directions": cell_dirs,
                           "Reference Cell Centroids": rcentroids,
                           "Cell IDs": cell_ids})

    return results


def parse_directory_file(filename: str) -> List[str]:
    with open(filename) as user:
        user_settings = yaml.load(user, yaml.SafeLoader)

    directories = {}
    for k, v in list(user_settings.items()):
        directories[k] = v

    return directories


def write_results(filename: str, results: dict) -> None:
    print(f"... Saving results to {filename}/results.xlsx")
    wb = Workbook()
    ws = []
    for i, k in enumerate(results.keys()):
        if i == 0:
            ws.append(wb.active)
            ws[-1].title = os.path.basename(k)
        else:
            ws.append(wb.create_sheet(title=os.path.basename(k)))
        ws[-1].append(['Cell ID',
                       'Volumetric Strain',
                       'Normalized Surface Area Change',
                       'Maximum Tensile Strain',
                       'Maximum Compressive Strain',
                       'Max Tensile Strain Component 1',
                       'Max Tensile Strain Component 2',
                       'Max Tensile Strain Component 3',
                       'Max Compressive Strain Component 1',
                       'Max Compressive Strain Component 2',
                       'Max Compressive Strain Component 3',
                       'Reference Volume',
                       'Deformed Volume',
                       'Reference Surface Area',
                       'Deformed Surface Area',
                       'Reference Surface Area to Volume',
                       'Deformed Surface Area to Volume',
                       'Reference Cell Direction Component 1',
                       'Reference Cell Direction Component 2',
                       'Reference Cell Direction Component 3',
                       'Reference Cell Centroid X',
                       'Reference Cell Centroid Y',
                       'Reference Cell Centroid Z',
                       'Coherence to Maximum Tensile Strain',
                       'Coherence to Maximum Compressive Strain'])

        for j, v in enumerate(results[k]['Volumetric Strains']):
            t_coherence = np.abs(np.dot(results[k]['Reference Cell Directions'][j],
                                        results[k]['1st Principal Strain Directions'][j]))
            c_coherence = np.abs(np.dot(results[k]['Reference Cell Directions'][j],
                                        results[k]['3rd Principal Strain Directions'][j]))
            ws[i].append([results[k]['Cell IDs'][j],
                          v,
                          results[k]['Normalized Surface Area Change'][j],
                          results[k]['1st Principal Strains'][j],
                          results[k]['3rd Principal Strains'][j],
                          results[k]['1st Principal Strain Directions'][j][0],
                          results[k]['1st Principal Strain Directions'][j][1],
                          results[k]['1st Principal Strain Directions'][j][2],
                          results[k]['3rd Principal Strain Directions'][j][0],
                          results[k]['3rd Principal Strain Directions'][j][1],
                          results[k]['3rd Principal Strain Directions'][j][2],
                          results[k]['Reference Volumes'][j],
                          results[k]['Deformed Volumes'][j],
                          results[k]['Reference Surface Areas'][j],
                          results[k]['Deformed Surface Areas'][j],
                          results[k]['Reference Surface Areas'][j] / results[k]['Reference Volumes'][j],
                          results[k]['Deformed Surface Areas'][j] / results[k]['Deformed Volumes'][j],
                          results[k]['Reference Cell Directions'][j][0],
                          results[k]['Reference Cell Directions'][j][1],
                          results[k]['Reference Cell Directions'][j][2],
                          results[k]['Reference Cell Centroids'][j][0],
                          results[k]['Reference Cell Centroids'][j][1],
                          results[k]['Reference Cell Centroids'][j][2],
                          t_coherence,
                          c_coherence]
                         )

    wb.save(filename=os.path.join(filename, 'results.xlsx'))


def main(directory_file):
    directories = parse_directory_file(directory_file)
    results = OrderedDict()

    for deformed_directory in directories['deformed']:
        rsurfs, dsurfs, cell_ids = read_surfaces(directories['reference'], deformed_directory)
        results[deformed_directory] = get_strain(rsurfs, dsurfs, cell_ids)

    write_results(directories['reference'], results)


if __name__ == '__main__':
    main(sys.argv[-1])
